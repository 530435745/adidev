from filter.workers.base import AdvancedWorkerBase
from filter.utils.xlsx_to_rows import xlsx_to_rows
from filter.config import *
from datetime import datetime
import openpyxl
import requests
import time
import os
import json


class F3Worker(AdvancedWorkerBase):
    TARGET_INFO = {}
    rows = xlsx_to_rows(GLOBAL_TARGET_RULES_FILE)
    for row in rows[1:]:
        TARGET_INFO[row[0]] = [row[1], row[0], row[2], row[3], row[4]]

    RESULTS = {}
    for customer, file_name in MATCH_RULES_FILES.items():
        RESULTS[customer] = {}
        rows = xlsx_to_rows(file_name)
        for row in rows[1:]:
            RESULTS[customer][f"{row[0]}-{row[1]}"] = {
                "id": row[2],
                "current": row[3],
                "name": row[4],
                "code": row[5],
                "type": row[6],
                "province": row[7],
                "address": row[8],
                "city": row[9]
            }

    def get_output_files(self):
        return [self.input_file.replace("F2_", "F3_")]

    def get_backup_file(self):
        return os.path.join(
            os.path.split(self.input_file)[0],
            "status",
            f"done{datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}_{os.path.split(self.input_file)[1]}"
        )

    @classmethod
    def set_rule(cls, customer):
        cls.RESULTS[customer] = {}
        wb = openpyxl.Workbook()
        ws = wb.active
        title = ["originCustomerName", "reference", "别名ID", "当前状态",
                 "customerName", "customerCode", "customerType", "customerProvince",
                 "customerAddress", "customerCity"]
        data = [title]
        for index_i, i in enumerate(data):
            for index_j, j in enumerate(i):
                ws.cell(index_i + 1, index_j + 1).value = j
        wb.save(os.path.join(ORDERED_FILES_DIR, customer, "customer_list.xlsx"))

    @classmethod
    def write_back(cls, customer, name, reference):
        back_file = os.path.join(ORDERED_FILES_DIR, customer, "customer_list.xlsx")
        to_add = [
            name, reference,
            cls.RESULTS[customer][f"{name}-{reference}"]["id"],
            cls.RESULTS[customer][f"{name}-{reference}"]["current"],
            cls.RESULTS[customer][f"{name}-{reference}"]["name"],
            cls.RESULTS[customer][f"{name}-{reference}"]["code"],
            cls.RESULTS[customer][f"{name}-{reference}"]["type"],
            cls.RESULTS[customer][f"{name}-{reference}"]["province"],
            cls.RESULTS[customer][f"{name}-{reference}"]["address"],
            cls.RESULTS[customer][f"{name}-{reference}"]["city"]
        ]
        wb = openpyxl.load_workbook(back_file)
        ws = wb.worksheets[0]
        data = []
        for i in ws.rows:
            if not i[0].value:
                break
            data.append([str(j.value).strip() if j.value is not None else "" for j in i])
        for index, i in enumerate(data):
            if i[0] == name and i[1] == reference:
                data[index] = to_add
                break
        else:
            data.append(to_add)
        for index_i, i in enumerate(data):
            for index_j, j in enumerate(i):
                ws.cell(index_i + 1, index_j + 1).value = j
        wb.save(back_file)

    @staticmethod
    def graceful_get(url, data_dict):
        while True:
            try:
                response = requests.get(url, params=data_dict)
                json_data = response.json()
                return json_data
            except Exception as e:
                print("Connection failed.", e)
                time.sleep(1)

    @classmethod
    def get_from_online(cls, customer, name, reference):
        print(f"Trying to get {name}-{reference} from server.")
        if info := cls.RESULTS[customer].get(f"{name}-{reference}"):
            json_data = cls.graceful_get(GET_RESULT_URL, {"case_id": info["id"]})
            if not json_data["data"]:
                return json_data["message"]
        else:
            json_data = cls.graceful_get(NEW_QUERY_URL, {"query": name, "reference": reference})
            if not json_data["data"]:
                return json_data["message"]
            case_id = json_data["data"]["case_id"]
            time.sleep(1)
            json_data = cls.graceful_get(GET_RESULT_URL, {"case_id": case_id})
            if not json_data["data"]:
                return json_data["message"]
        cls.RESULTS[customer][f"{name}-{reference}"] = {
            "id": json_data["data"]["case_id"],
            "current": json_data["data"]["status"]
        }
        business = json_data["data"].get("business")
        if not business:
            business = {}
        cls.RESULTS[customer][f"{name}-{reference}"].update({
            "name": business.get("name"),
            "code": business.get("p_code"),
            "type": business.get("ext3"),
            "province": business.get("province"),
            "address": business.get("addr"),
            "city": business.get("city")
        })
        cls.write_back(customer, name, reference)

    def real_process(self):
        print(f"F3: {self.input_file}")
        if self.file_type == "I":
            return True
        if self.customer not in self.RESULTS:
            self.set_rule(self.customer)
        dealer_indexes = list(
            map(self.data[0].index,
                ["dealerName", "dealerCode", "dealerProvince", "dealerCity", "dealerLevel"]))
        if self.file_type == "S":
            names = ["customerName", "customerCode", "customerType", "customerProvince",
                    "customerAddress", "customerCity"]
        else:
            names = ["supplierName", "supplierCode"]
        data_indexes = list(map(self.data[0].index, names))
        reference_index = self.data[0].index("dealerProvince")
        repeat_data = self.TARGET_INFO.get(self.customer)
        if not repeat_data:
            self.error(f"未在目标清单中找到经销商信息，经销商代码：{self.customer}")
            return False
        errors = {}
        for row_index, row in enumerate(self.data):
            if row_index == 0:
                row.append("origin" + names[0][0].upper() + names[0][1:])
                continue
            for index in range(5):
                row[dealer_indexes[index]] = repeat_data[index]
            if not row[data_indexes[0]]:
                self.error(f"第{row_index + 1}行中，supplierName或customerName列为空值，无法继续进行匹配")
                return False
            if f"{row[data_indexes[0]]}-{row[reference_index]}" in errors:
                continue
            failed = False
            while True:
                if exist_info := self.RESULTS[self.customer].get(f"{row[data_indexes[0]]}-{row[reference_index]}"):
                    if exist_info["current"] in ["E1", "W1", "M1"]:
                        row.append(row[data_indexes[0]])
                        row[data_indexes[0]] = exist_info["name"]
                        row[data_indexes[1]] = exist_info["code"]
                        if self.file_type == "S":
                            row[data_indexes[2]] = exist_info["type"]
                            row[data_indexes[3]] = exist_info["province"]
                            row[data_indexes[4]] = exist_info["address"]
                            row[data_indexes[5]] = exist_info["city"]
                        break
                    else:
                        if failed:
                            errors[f"{row[data_indexes[0]]}-{row[reference_index]}"] = \
                                f"第{row_index + 1}行查询无结果，查询信息：{{\"query\": \"{row[data_indexes[0]]}\", \"reference\": \"{row[reference_index]}\"}}"
                            break
                        else:
                            msg = self.get_from_online(
                                self.customer, row[data_indexes[0]], row[reference_index]
                            )
                            if msg:
                                errors[f"{row[data_indexes[0]]}-{row[reference_index]}"] = \
                                    f"第{row_index + 1}行查询出错，查询信息：{{\"query\": \"{row[data_indexes[0]]}\", \"reference\": \"{row[reference_index]}\"}}，返回信息：{msg}"
                                break
                            failed = True
                            continue
                else:
                    msg = self.get_from_online(self.customer, row[data_indexes[0]], row[reference_index])
                    if msg:
                        errors[f"{row[data_indexes[0]]}-{row[reference_index]}"] = \
                            f"第{row_index + 1}行查询出错，查询信息：{{\"query\": \"{row[data_indexes[0]]}\", \"reference\": \"{row[reference_index]}\"}}，返回信息：{msg}"
                        break
                    failed = True
                    continue
        if errors:
            self.error("\n".join(errors.values()))
            return False
        return True

from workers.base import AdvancedWorkerBase
from utils.xlsx_to_rows import xlsx_to_rows
from config import *
from datetime import datetime
import openpyxl
import requests
import time
import os


class F3Worker(AdvancedWorkerBase):
    TARGET_INFO = {}
    rows = xlsx_to_rows(GLOBAL_TARGET_RULES_FILE)
    for row in rows[1:]:
        TARGET_INFO[row[0].split("_")[1]] = [row[1], row[0], row[2], row[3], row[4]]

    RESULTS = {}
    for customer, file_name in MATCH_RULES_FILES.items():
        RESULTS[customer] = {}
        rows = xlsx_to_rows(file_name)
        for row in rows[1:]:
            RESULTS[customer][f"{row[0]}-{row[1]}"] = {
                "id": row[2],
                "current": row[3],
                "name": row[4],
                "code": row[5],
                "type": row[6],
                "province": row[7],
                "address": row[8],
                "city": row[9]
            }

    def get_output_files(self):
        return [self.input_file.replace("F2_", "F3_")]

    def get_backup_file(self):
        return os.path.join(
            os.path.split(self.input_file)[0],
            "status",
            f"done{datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}_{os.path.split(self.input_file)[1]}"
        )

    @classmethod
    def set_rule(cls, factory, customer):
        cls.RESULTS[customer] = {}
        wb = openpyxl.Workbook()
        ws = wb.active
        title = ["originCustomerName", "reference", "别名ID", "当前状态",
                 "customerName", "customerCode", "customerProvince",
                 "customerAddress", "customerCity"]
        data = [title]
        for index_i, i in enumerate(data):
            for index_j, j in enumerate(i):
                ws.cell(index_i + 1, index_j + 1).value = j
        wb.save(os.path.join(ORDERED_FILES_DIR, f"{factory}_{customer}", "customer_list.xlsx"))

    @classmethod
    def write_back(cls, factory, customer, name, reference):
        back_file = os.path.join(ORDERED_FILES_DIR, f"{factory}_{customer}", "customer_list.xlsx")
        to_add = [
            name, reference, cls.RESULTS[customer][f"{name}-{reference}"]["id"], cls.RESULTS[customer][f"{name}-{reference}"]["current"],
            cls.RESULTS[customer][f"{name}-{reference}"]["name"], cls.RESULTS[customer][f"{name}-{reference}"]["code"], cls.RESULTS[customer][f"{name}-{reference}"]["type"],
            cls.RESULTS[customer][f"{name}-{reference}"]["province"], cls.RESULTS[customer][f"{name}-{reference}"]["address"], cls.RESULTS[customer][f"{name}-{reference}"]["city"]
        ]
        wb = openpyxl.load_workbook(back_file)
        ws = wb.worksheets[0]
        data = []
        for i in ws.rows:
            if not i[0].value:
                break
            data.append([str(j.value).strip() if j.value is not None else "" for j in i])
        for index, i in enumerate(data):
            if i[0] == name and i[1] == reference:
                data[index] = to_add
                break
        else:
            data.append(to_add)
        for index_i, i in enumerate(data):
            for index_j, j in enumerate(i):
                ws.cell(index_i + 1, index_j + 1).value = j
        wb.save(back_file)

    @classmethod
    def get_from_online(cls, factory, customer, name, reference):
        print(f"Trying to get {name}-{reference} from server.")
        if info := cls.RESULTS[customer].get(f"{name}-{reference}"):
            response = requests.get(GET_RESULT_URL, params={"case_id": info["id"]})
            json_data = response.json()
        else:
            response = requests.get(NEW_QUERY_URL, params={"query": name, "reference": reference})
            json_data = response.json()
            case_id = json_data["data"]["case_id"]
            time.sleep(3)
            response = requests.get(GET_RESULT_URL, params={"case_id": case_id})
            json_data = response.json()
        cls.RESULTS[customer][f"{name}-{reference}"] = {
            "id": json_data["data"]["case_id"],
            "current": json_data["data"]["status"]
        }
        business = json_data["data"].get("business", {})
        cls.RESULTS[customer][f"{name}-{reference}"].update({
            "name": business.get("name"),
            "code": business.get("p_code"),
            "type": business.get("ext3"),
            "province": business.get("province"),
            "address": business.get("addr"),
            "city": business.get("city")
        })
        cls.write_back(factory, customer, name, reference)

    def real_process(self):
        print(f"F3: {self.input_file}")
        if self.customer not in self.RESULTS:
            self.set_rule(self.factory_code, self.customer)
        dealer_indexes = list(
            map(self.data[0].index ,
                ["dealerName", "dealerCode", "dealerProvince", "dealerCity", "dealerLevel"]))
        if self.file_type == "S":
            name_index = self.data[0].index("customerName")
            data_indexes = list(
                map(self.data[0].index, ["customerName", "customerCode", "customerType", "customerProvince",
                                         "customerAddress", "customerCity"]))
        else:
            name_index = self.data[0].index("supplierName")
            data_indexes = list(
                map(self.data[0].index, ["supplierName", "supplierCode"]))
        reference_index = self.data[0].index("dealerProvince")
        repeat_data = self.TARGET_INFO.get(self.customer)
        if not repeat_data:
            self.error(f"当前文件：{self.input_file}, 未在目标清单中找到经销商信息，经销商代码：{self.customer}")
            return False
        for row_index, row in enumerate(self.data[1:]):
            for index in range(5):
                row[dealer_indexes[index]] = repeat_data[index]
            failed = False
            while True:
                if exist_info := self.RESULTS[self.customer].get(f"{row[name_index]}-{row[reference_index]}"):
                    if exist_info["current"] in ["E1", "W1", "M1"]:
                        row[data_indexes[0]] = exist_info["name"]
                        row[data_indexes[1]] = exist_info["code"]
                        if self.file_type == "S":
                            row[data_indexes[2]] = exist_info["type"]
                            row[data_indexes[3]] = exist_info["province"]
                            row[data_indexes[4]] = exist_info["address"]
                            row[data_indexes[5]] = exist_info["city"]
                        break
                    else:
                        if failed:
                            self.error(f"文件{self.input_file}第{row_index}行查询无结果，"
                                       f"查询信息：{{'query': '{row[name_index]}', 'reference': '{row[reference_index]}'}}")
                            return False
                        else:
                            self.get_from_online(self.factory_code, self.customer, row[name_index], row[reference_index])
                            failed = True
                            continue
                else:
                    self.get_from_online(self.factory_code, self.customer, row[name_index], row[reference_index])
                    failed = True
                    continue
        return True




